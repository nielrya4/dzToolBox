"""
Celery tasks for long-running operations
"""

from celery_app import celery_app
from utils import spreadsheet, compression, embedding
from utils.output import Output
from utils.project import project_from_json
from server import database
from dz_lib.univariate import distributions
from dz_lib.utils import data
import secrets
import os
import sys
import traceback


def clean_sample_name(sample_name):
    """Clean sample names (convert floats to ints if possible)"""
    try:
        num = float(sample_name)
        if num.is_integer():
            return str(int(num))
    except ValueError:
        pass
    return str(sample_name)


@celery_app.task(bind=True)
def tensor_factorization_task(
    self,
    project_id,
    user_id,
    output_title,
    sample_names,
    rank,
    model_type,
    output_types,
    normalization_method='standardize',
    padding_mode='zero',
    font_name='Ubuntu',
    font_size=12,
    fig_width=10,
    fig_height=8,
    color_map='viridis',
    stack_graphs='true',
    fill='false'
):
    """
    Run multivariate tensor factorization as a Celery task

    Parameters match the route arguments but are serialized
    """
    try:
        # Import tensor_factorization here to avoid Julia import at module load time
        from utils import tensor_factorization
        
        # Import Flask app inside function to avoid circular import
        from dzToolBox import app as flask_app

        # Update task state to show progress
        self.update_state(state='PROGRESS', meta={'status': 'Loading project...'})

        # Load project from database (needs Flask app context)
        with flask_app.app_context():
            # Import CodeFile model from dzToolBox module
            import dzToolBox as APP
            # Query database directly without relying on current_user
            file = APP.CodeFile.query.filter_by(user_id=user_id, id=project_id).first_or_404()
            project_content = compression.decompress(file.content)

        project = project_from_json(project_content)

        self.update_state(state='PROGRESS', meta={'status': 'Loading multivariate samples...'})

        # Load multivariate samples from row-based data (use grainalyzer_data)
        spreadsheet_data = spreadsheet.text_to_array(project.grainalyzer_data)

        # Convert to row format if needed (assuming data might be stored in old format)
        # For now, we assume the project data is already in row format
        # If it's in column format, we'd need to handle that differently
        loaded_samples, feature_names = spreadsheet.read_multivariate_samples(
            spreadsheet_array=spreadsheet_data,
            max_age=4500
        )

        # Filter to selected samples
        active_samples = []
        for sample in loaded_samples:
            # Clean sample name
            clean_name = clean_sample_name(sample.name)
            sample.name = clean_name
            if clean_name in sample_names:
                active_samples.append(sample)

        if len(active_samples) < 2:
            raise ValueError(f"At least 2 samples required, got {len(active_samples)}")

        self.update_state(state='PROGRESS', meta={'status': 'Running factorization using original dzgrainalyzer code...'})

        # Call original Julia code directly to match dzgrainalyzer exactly
        # This uses rank_sources_custom_rank() from dzgrainalyzer_helpers.jl
        import subprocess
        import tempfile
        import json

        temp_excel = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_json = tempfile.NamedTemporaryFile(suffix='.json', delete=False, mode='w')
        temp_excel.close()
        temp_json.close()

        try:
            # Transform to column-based format expected by read_raw_data()
            # One sheet per feature, columns = samples, rows = grains
            import openpyxl
            wb = openpyxl.Workbook()

            # Find max grain count
            max_grains = max([len(sample.grains) for sample in active_samples])

            # Create one sheet per feature
            for feat_idx, feature_name in enumerate(feature_names):
                if feat_idx == 0:
                    ws = wb.active
                    ws.title = feature_name
                else:
                    ws = wb.create_sheet(title=feature_name)

                # Write data: rows = grains, columns = samples (no headers)
                for grain_idx in range(max_grains):
                    row = []
                    for sample in active_samples:
                        if grain_idx < len(sample.grains):
                            val = sample.grains[grain_idx].features.get(feature_name, None)
                            row.append(val)
                        else:
                            row.append(None)  # Pad with None for missing grains
                    ws.append(row)

            wb.save(temp_excel.name)
            print(f"Temporary Excel created (column-based format): {temp_excel.name}")

            # Get Julia executable path from juliapkg
            import juliapkg
            julia_exe = juliapkg.executable()
            julia_project = juliapkg.project()

            # Call Julia script with environment set to use juliacall packages
            julia_script = os.path.join(os.path.dirname(__file__), 'julia_scripts', 'run_factorization.jl')
            cmd = [julia_exe, '--project=' + julia_project, julia_script, temp_excel.name, str(rank), temp_json.name]

            print(f"Running Julia command: {' '.join(cmd)}")
            print(f"Julia project: {julia_project}")
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=600)

            print(f"Julia stdout: {result.stdout}")
            if result.stderr:
                print(f"Julia stderr: {result.stderr}")

            if result.returncode != 0:
                raise RuntimeError(f"Julia script failed with code {result.returncode}: {result.stderr}")

            # Read JSON results
            with open(temp_json.name, 'r') as f:
                julia_results = json.load(f)

            if julia_results.get('status') == 'error':
                raise RuntimeError(f"Julia script error: {julia_results.get('error')}")

            print(f"Julia factorization complete")

        finally:
            # Clean up temp files
            if os.path.exists(temp_excel.name):
                os.unlink(temp_excel.name)
            if os.path.exists(temp_json.name):
                os.unlink(temp_json.name)

        # Extract data from Julia results
        # Julia returns:
        # - measurement_data: empirical KDE visualization data
        # - learned_densities: learned source KDE data
        # - learned_coefficients: C matrix coefficient data
        # - source_identification_per_sink: grain-level source attribution

        sample_names_list = [s.replace("sink ", "") for s in julia_results.get('sinks', [])]
        feature_names_from_julia = [md['name'] for md in julia_results.get('measurement_data', [])]
        r2 = 0.0  # Julia doesn't return R² in this function

        self.update_state(state='PROGRESS', meta={'status': 'Generating visualizations...'})

        pending_outputs = []

        # Debug: log which output types were requested and Julia results
        print(f"DEBUG: Generating visualizations. output_types = {output_types}")
        print(f"DEBUG: Sample names from Julia: {sample_names_list}")
        print(f"DEBUG: Features from Julia: {feature_names_from_julia}")
        print(f"DEBUG: Julia results keys: {julia_results.keys()}")

        print(f"Factorization completed successfully using original dzgrainalyzer code")
        print(f"Rank: {rank}, Samples: {len(sample_names_list)}, Features: {len(feature_names_from_julia)}")

        # Generate source attribution scatter plots if requested
        if "source_attribution_plots" in output_types:
            print(f"DEBUG: Generating source attribution plots from Julia data...")

            # Transform Julia format to Python visualization format
            # Julia: {name: "sink 1", data: {sources: [1,2,1,...], loglikelihood_ratios: [0.9,...]}}
            # Python: {sample_name: "1", grain_attributions: [1,2,1,...], grain_confidences: [0.9,...]}
            # Note: Keep 1-indexed sources to match visualization Y-axis ticks

            attribution_results = []
            all_sources_seen = set()
            for sink_data in julia_results.get('source_identification_per_sink', []):
                sample_name = sink_data['name'].replace('sink ', '')  # "sink 1" -> "1"
                sources = sink_data['data']['sources']
                confidences = sink_data['data']['loglikelihood_ratios']

                # DEBUG: Check what sources Julia is returning
                unique_sources = set(sources)
                all_sources_seen.update(unique_sources)
                print(f"DEBUG: Sample {sample_name} has {len(sources)} grains, sources: {sorted(unique_sources)} (Julia 1-indexed)")
                if len(sources) >= 5:
                    print(f"DEBUG: Sample {sample_name} first 5 sources: {sources[:5]}")

                # Keep Julia's 1-indexed sources (don't convert to 0-indexed)
                # The visualization expects 1-indexed sources to match the Y-axis ticks
                grain_attributions = sources

                # Convert confidences to floats and handle 'null' values
                # Julia's clean_inf() converts -Inf to "null" string
                grain_confidences = []
                for conf in confidences:
                    if conf == 'null' or conf is None:
                        grain_confidences.append(0.0)  # Low confidence for null values
                    else:
                        # Ensure it's a float, even if JSON parsed it as string
                        grain_confidences.append(float(conf))

                attribution_results.append({
                    'sample_name': sample_name,
                    'grain_attributions': grain_attributions,
                    'grain_confidences': grain_confidences
                })

            print(f"DEBUG: ALL sources seen across all samples (Julia 1-indexed): {sorted(all_sources_seen)}")
            print(f"DEBUG: Expected rank={rank} sources, so should see sources 1 through {rank}")

            print(f"DEBUG: Transformed {len(attribution_results)} samples for visualization")

            # Visualize source attribution for each sample (returns list of tabs)
            tabs = tensor_factorization.visualize_source_attribution_tabbed(
                attributions=attribution_results,
                rank=rank,
                title=f"{output_title}",
                font_path=f'static/global/fonts/{font_name}.ttf',
                font_size=font_size,
                fig_width=fig_width,
                fig_height=fig_height,
                color_map='Greens'
            )

            output_id = secrets.token_hex(15)
            output_data = embedding.embed_tabbed_graphs(
                tabs=tabs,
                output_id=output_id,
                project_id=project_id,
                fig_type="matplotlib",
                img_format='svg',
                download_formats=['svg', 'png'],
                is_grainalyzer=True
            )
            pending_outputs.append({
                "output_id": output_id,
                "output_type": "tabbed_graph",
                "output_data": output_data
            })
            print(f"DEBUG: Source attribution scatter plots added. Total outputs: {len(pending_outputs)}")

        # Generate source attribution matrix if requested
        if "source_attribution_matrix" in output_types:
            print(f"DEBUG: Generating source attribution matrix from Julia data...")

            # Create tabbed matrix output for source attribution (exportable as XLSX)
            import pandas as pd
            matrix_tabs = []
            for sink_data in julia_results.get('source_identification_per_sink', []):
                sample_name = sink_data['name'].replace('sink ', '')
                sources = sink_data['data']['sources']
                confidences = sink_data['data']['loglikelihood_ratios']

                # Convert confidences to floats
                clean_confidences = []
                for conf in confidences:
                    if conf == 'null' or conf is None:
                        clean_confidences.append(0.0)
                    else:
                        clean_confidences.append(float(conf))

                # Create DataFrame for this sample
                df = pd.DataFrame({
                    'loglikelihood_ratios': clean_confidences,
                    'sources': sources
                })

                matrix_tabs.append({
                    'name': sample_name,
                    'dataframe': df
                })

            # Embed as tabbed matrix
            output_id = secrets.token_hex(15)
            output_data = embedding.embed_tabbed_matrices(
                tabs=matrix_tabs,
                output_id=output_id,
                project_id=project_id,
                download_formats=['xlsx', 'csv'],
                is_grainalyzer=True
            )
            pending_outputs.append({
                "output_id": output_id,
                "output_type": "tabbed_matrix",
                "output_data": output_data
            })
            print(f"DEBUG: Source attribution matrix added. Total outputs: {len(pending_outputs)}")

        # Generate learned source KDEs visualization
        if "learned_source_kdes" in output_types:
            print(f"DEBUG: Generating learned source KDEs from Julia data...")

            # Julia format: [{name: "Age", data: [{domain: 100, "source 1": 0.01, "source 2": 0.02}, ...]}, ...]
            learned_densities = julia_results.get('learned_densities', [])

            # Use dedicated visualization function (matches empirical KDEs style)
            # stack_sources = True means overlay (like stack_samples=True for empirical)
            # stack_sources = False means stacked vertically (like stack_samples=False for empirical)
            stack_sources = stack_graphs != "true"
            fill_kdes = fill == "true"

            figures = tensor_factorization.visualize_learned_source_kdes_from_julia(
                learned_densities=learned_densities,
                rank=rank,
                title=f"{output_title}",
                font_path=f'static/global/fonts/{font_name}.ttf',
                font_size=font_size,
                fig_width=fig_width,
                fig_height=fig_height,
                color_map=color_map,
                stack_sources=stack_sources,
                fill=fill_kdes
            )

            # Convert figures to tabs format
            tabs = []
            for i, fig in enumerate(figures):
                feature_name = learned_densities[i]['name']
                tabs.append({
                    'name': feature_name,
                    'fig': fig
                })

            output_id = secrets.token_hex(15)
            output_data = embedding.embed_tabbed_graphs(
                tabs=tabs,
                output_id=output_id,
                project_id=project_id,
                fig_type="matplotlib",
                img_format='svg',
                download_formats=['svg', 'png'],
                is_grainalyzer=True
            )
            pending_outputs.append({
                "output_id": output_id,
                "output_type": "tabbed_graph",
                "output_data": output_data
            })
            print(f"DEBUG: Learned source KDEs added. Total outputs: {len(pending_outputs)}")

        # Generate learned coefficients visualization
        if "learned_coefficients" in output_types:
            print(f"DEBUG: Generating learned coefficients from Julia data...")

            # Julia format: [{name: "sink 1", data: [0.3, 0.5, 0.2]}, ...]
            coefficients_data = julia_results.get('learned_coefficients', [])

            # Create horizontal stacked bar chart
            import matplotlib.pyplot as plt
            import numpy as np

            fig, ax = plt.subplots(1, 1, figsize=(fig_width, fig_height))

            # Build data
            y_pos = np.arange(len(sample_names_list))
            left_offset = np.zeros(len(sample_names_list))

            for source_idx in range(rank):
                coeffs = [coef_dict['data'][source_idx] for coef_dict in coefficients_data]
                ax.barh(y_pos, coeffs, left=left_offset, label=f'source {source_idx + 1}')
                left_offset += np.array(coeffs)

            ax.set_yticks(y_pos)
            ax.set_yticklabels(sample_names_list, fontsize=font_size-2)
            ax.set_xlabel('coefficient', fontsize=font_size)
            ax.set_ylabel('sample', fontsize=font_size)
            ax.legend(fontsize=font_size-2, loc='center left', bbox_to_anchor=(1, 0.5))
            ax.grid(True, alpha=0.3, axis='x')
            plt.tight_layout()

            output_id = secrets.token_hex(15)
            output_data = embedding.embed_graph(
                fig=fig,
                output_id=output_id,
                project_id=project_id,
                fig_type="matplotlib",
                img_format='svg',
                download_formats=['svg', 'png'],
                is_grainalyzer=True
            )
            pending_outputs.append({
                "output_id": output_id,
                "output_type": "graph",
                "output_data": output_data
            })
            print(f"DEBUG: Learned coefficients added. Total outputs: {len(pending_outputs)}")

        # Return outputs for preview (don't auto-save)
        print(f"DEBUG: Returning {len(pending_outputs)} outputs for preview")

        return {
            "status": "completed",
            "outputs": pending_outputs,  # Include full output_data for preview
            "saved": False,  # Not saved yet - will be saved via preview modal
            "r2": r2
        }

    except Exception as e:
        # Print detailed error to worker logs
        error_msg = str(e)
        error_trace = traceback.format_exc()
        print("=" * 80, file=sys.stderr)
        print(f"TENSOR FACTORIZATION ERROR:", file=sys.stderr)
        print(f"Error: {error_msg}", file=sys.stderr)
        print(f"Type: {type(e).__name__}", file=sys.stderr)
        print("Traceback:", file=sys.stderr)
        print(error_trace, file=sys.stderr)
        print("=" * 80, file=sys.stderr)
        sys.stderr.flush()

        # Re-raise the exception to let Celery handle it properly
        raise


@celery_app.task(bind=True)
def view_empirical_kdes_task(
    self,
    project_id,
    user_id,
    output_title,
    sample_names,
    font_name='Ubuntu',
    font_size=12,
    fig_width=10,
    fig_height=8,
    color_map='tab20'
):
    """
    View empirical KDEs without running factorization

    This is a lightweight task that just visualizes the raw input data.
    """
    try:
        # Import tensor_factorization here to avoid Julia import at module load time
        from utils import tensor_factorization
        
        from dzToolBox import app as flask_app

        self.update_state(state='PROGRESS', meta={'status': 'Loading project...'})

        # Load project from database
        with flask_app.app_context():
            import dzToolBox as APP
            file = APP.CodeFile.query.filter_by(user_id=user_id, id=project_id).first_or_404()
            project_content = compression.decompress(file.content)

        project = project_from_json(project_content)

        self.update_state(state='PROGRESS', meta={'status': 'Loading multivariate samples...'})

        # Load multivariate samples
        spreadsheet_data = spreadsheet.text_to_array(project.grainalyzer_data)
        loaded_samples, feature_names = spreadsheet.read_multivariate_samples(
            spreadsheet_array=spreadsheet_data,
            max_age=4500
        )

        # Filter to selected samples
        active_samples = []
        for sample in loaded_samples:
            clean_name = clean_sample_name(sample.name)
            sample.name = clean_name
            if clean_name in sample_names:
                active_samples.append(sample)

        if len(active_samples) < 2:
            raise ValueError(f"At least 2 samples required, got {len(active_samples)}")

        self.update_state(state='PROGRESS', meta={'status': 'Building tensor...'})

        # Create tensor from multivariate samples
        tensor, metadata = tensor_factorization.create_tensor_from_multivariate_samples(
            samples=active_samples,
            feature_names=feature_names,
            padding_mode='zero'
        )

        self.update_state(state='PROGRESS', meta={'status': 'Generating visualization...'})

        # Check if graphs should be stacked (from project settings)
        # stack_graphs = "true" means separate subplots stacked vertically
        # stack_graphs = "false" means overlaid on one plot
        stack_samples = project.settings.graph_settings.stack_graphs != "true"

        # Check if fill is enabled
        fill = project.settings.graph_settings.fill == "true"

        output_id = secrets.token_hex(15)

        # Generate empirical KDEs visualization (one figure per feature for tabs)
        # stack_samples=True means overlay, stack_samples=False means separate subplots
        feature_figures = tensor_factorization.visualize_empirical_kdes_tabbed(
            tensor=tensor,
            feature_names=metadata['feature_names'],
            sample_names=metadata['sample_names'],
            grain_counts=metadata['grain_counts'],
            title=f"{output_title}",
            font_path=f'static/global/fonts/{font_name}.ttf',
            font_size=font_size,
            fig_width=fig_width,
            fig_height=fig_height,
            color_map=color_map,
            stack_samples=stack_samples,
            fill=fill
        )

        # Create tabs - one per feature
        tabs = []
        for feature_name, fig in zip(metadata['feature_names'], feature_figures):
            tabs.append({
                "name": feature_name,
                "fig": fig
            })

        # Embed as tabbed output
        output_data = embedding.embed_tabbed_graphs(
            tabs=tabs,
            output_id=output_id,
            project_id=project_id,
            fig_type="matplotlib",
            img_format='svg',
            download_formats=['svg', 'png'],
            is_grainalyzer=True
        )

        pending_outputs = [{
            "output_id": output_id,
            "output_type": "tabbed_graph",
            "output_data": output_data
        }]

        # Return outputs for preview (don't auto-save)
        return {
            "status": "completed",
            "outputs": pending_outputs,  # Include full output_data for preview
            "saved": False  # Not saved yet - will be saved via preview modal
        }

    except Exception as e:
        error_msg = str(e)
        error_trace = traceback.format_exc()
        print("=" * 80, file=sys.stderr)
        print(f"VIEW EMPIRICAL KDES ERROR:", file=sys.stderr)
        print(f"Error: {error_msg}", file=sys.stderr)
        print(f"Type: {type(e).__name__}", file=sys.stderr)
        print("Traceback:", file=sys.stderr)
        print(error_trace, file=sys.stderr)
        print("=" * 80, file=sys.stderr)
        sys.stderr.flush()

        # Re-raise the exception to let Celery handle it properly
        raise


@celery_app.task(bind=True)
def find_optimal_rank_task(
    self,
    project_id,
    user_id,
    output_title,
    sample_names,
    min_rank,
    max_rank,
    model_type,
    output_types,
    normalization_method='standardize',
    font_name='Ubuntu',
    font_size=12,
    fig_width=10,
    fig_height=8,
    color_map='viridis'
):
    """
    Find optimal rank by testing multiple ranks

    This runs factorization at each rank in the range and generates
    the rank selection visualization (misfit vs ranks + optimal rank).
    """
    try:
        # Import tensor_factorization here to avoid Julia import at module load time
        from utils import tensor_factorization
        
        from dzToolBox import app as flask_app

        self.update_state(state='PROGRESS', meta={'status': 'Loading project...'})

        # Load project from database
        with flask_app.app_context():
            import dzToolBox as APP
            file = APP.CodeFile.query.filter_by(user_id=user_id, id=project_id).first_or_404()
            project_content = compression.decompress(file.content)

        project = project_from_json(project_content)

        self.update_state(state='PROGRESS', meta={'status': 'Loading multivariate samples...'})

        # Load multivariate samples
        spreadsheet_data = spreadsheet.text_to_array(project.grainalyzer_data)
        loaded_samples, feature_names = spreadsheet.read_multivariate_samples(
            spreadsheet_array=spreadsheet_data,
            max_age=4500
        )

        # Filter to selected samples
        active_samples = []
        for sample in loaded_samples:
            clean_name = clean_sample_name(sample.name)
            sample.name = clean_name
            if clean_name in sample_names:
                active_samples.append(sample)

        if len(active_samples) < 2:
            raise ValueError(f"At least 2 samples required, got {len(active_samples)}")

        self.update_state(state='PROGRESS', meta={'status': 'Calling Julia rank selection code...'})

        # Call original Julia code directly instead of reimplementing
        import subprocess
        import tempfile
        import json

        # Create temporary Excel file with row-based format expected by Julia
        temp_excel = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_json = tempfile.NamedTemporaryFile(suffix='.json', delete=False, mode='w')
        temp_excel.close()
        temp_json.close()

        try:
            # Transform to column-based format expected by read_raw_data()
            # One sheet per feature, columns = samples, rows = grains
            import openpyxl
            wb = openpyxl.Workbook()

            # Find max grain count
            max_grains = max([len(sample.grains) for sample in active_samples])

            # Create one sheet per feature
            for feat_idx, feature_name in enumerate(feature_names):
                if feat_idx == 0:
                    ws = wb.active
                    ws.title = feature_name
                else:
                    ws = wb.create_sheet(title=feature_name)

                # Write data: rows = grains, columns = samples (no headers)
                for grain_idx in range(max_grains):
                    row = []
                    for sample in active_samples:
                        if grain_idx < len(sample.grains):
                            val = sample.grains[grain_idx].features.get(feature_name, None)
                            row.append(val)
                        else:
                            row.append(None)  # Pad with None for missing grains
                    ws.append(row)

            wb.save(temp_excel.name)
            print(f"Temporary Excel created (column-based format): {temp_excel.name}")

            # Get Julia executable path from juliapkg
            import juliapkg
            julia_exe = juliapkg.executable()
            julia_project = juliapkg.project()

            # Call Julia script with environment set to use juliacall packages
            julia_script = os.path.join(os.path.dirname(__file__), 'julia_scripts', 'rank_selection.jl')
            cmd = [julia_exe, '--project=' + julia_project, julia_script, temp_excel.name, str(min_rank), str(max_rank), temp_json.name]

            print(f"Running Julia command: {' '.join(cmd)}")
            print(f"Julia project: {julia_project}")
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=600)

            print(f"Julia stdout: {result.stdout}")
            if result.stderr:
                print(f"Julia stderr: {result.stderr}")

            if result.returncode != 0:
                raise RuntimeError(f"Julia script failed with code {result.returncode}: {result.stderr}")

            # Read JSON results
            with open(temp_json.name, 'r') as f:
                julia_results = json.load(f)

            if julia_results.get('status') == 'error':
                raise RuntimeError(f"Julia script error: {julia_results.get('error')}")

            ranks = julia_results['ranks']
            errors = julia_results['relative_errors']
            curvatures = julia_results['curvatures']
            best_rank = julia_results['best_rank']
            max_r2 = julia_results.get('r2', 0.0)

            print(f"Julia analysis complete: best rank = {best_rank}")

        finally:
            # Clean up temp files
            if os.path.exists(temp_excel.name):
                os.unlink(temp_excel.name)
            if os.path.exists(temp_json.name):
                os.unlink(temp_json.name)

        self.update_state(state='PROGRESS', meta={'status': 'Creating visualizations...'})

        # Julia already calculated curvatures and selected best rank
        print(f"Rank selection: Best rank is {best_rank} (from Julia analysis)")
        print(f"Generating outputs: {output_types}")

        pending_outputs = []

        # Generate visualization outputs based on user selection
        # Users can request either or both plots via output_types parameter

        # Generate misfit plot if requested
        if "misfit_plot" in output_types:
            graph_fig = tensor_factorization.visualize_misfit_plot(
                ranks=ranks,
                errors=errors,
                title=f"{output_title}",
                font_path=f'static/global/fonts/{font_name}.ttf',
                font_size=font_size,
                fig_width=fig_width,
                fig_height=fig_height
            )

            output_id = secrets.token_hex(15)
            output_data = embedding.embed_graph(
                fig=graph_fig,
                output_id=output_id,
                project_id=project_id,
                fig_type="matplotlib",
                img_format='svg',
                download_formats=['svg', 'png'],
                is_grainalyzer=True
            )

            pending_outputs.append({
                "output_id": output_id,
                "output_type": "graph",
                "output_data": output_data
            })

        # Generate curvature plot if requested
        if "curvature_plot" in output_types:
            graph_fig = tensor_factorization.visualize_curvature_plot(
                ranks=ranks,
                curvatures=curvatures,
                best_rank=best_rank,
                title=f"{output_title}\nOptimum Rank (R²={max_r2:.3f})",
                font_path=f'static/global/fonts/{font_name}.ttf',
                font_size=font_size,
                fig_width=fig_width,
                fig_height=fig_height
            )

            output_id = secrets.token_hex(15)
            output_data = embedding.embed_graph(
                fig=graph_fig,
                output_id=output_id,
                project_id=project_id,
                fig_type="matplotlib",
                img_format='svg',
                download_formats=['svg', 'png'],
                is_grainalyzer=True
            )

            pending_outputs.append({
                "output_id": output_id,
                "output_type": "graph",
                "output_data": output_data
            })

        # Return outputs for preview (don't auto-save)
        return {
            "status": "completed",
            "outputs": pending_outputs,  # Include full output_data for preview
            "saved": False,  # Not saved yet
            "best_rank": best_rank,
            "max_r2": max_r2
        }

    except Exception as e:
        error_msg = str(e)
        error_trace = traceback.format_exc()
        print("=" * 80, file=sys.stderr)
        print(f"FIND OPTIMAL RANK ERROR:", file=sys.stderr)
        print(f"Error: {error_msg}", file=sys.stderr)
        print(f"Type: {type(e).__name__}", file=sys.stderr)
        print("Traceback:", file=sys.stderr)
        print(error_trace, file=sys.stderr)
        print("=" * 80, file=sys.stderr)
        sys.stderr.flush()

        # Re-raise the exception to let Celery handle it properly
        raise
