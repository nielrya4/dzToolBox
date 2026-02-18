from pandas import DataFrame
from dz_lib.utils import encode, formats, matrices
from dz_lib.utils.encode import get_mime_type, safe_filename
import secrets


def embed_graph(fig, output_id, project_id, fig_type="matplotlib", img_format='svg', download_formats=['svg'], is_grainalyzer=False):
    accepted_image_formats = ['jpg', 'jpeg', 'png', 'pdf', 'eps', 'svg']
    if not formats.check(file_format=img_format, accepted_formats=accepted_image_formats):
        raise ValueError("Image format is not supported.")
    for download_format in download_formats:
        if not formats.check(file_format=download_format, accepted_formats=accepted_image_formats):
            raise ValueError("Download format is not supported.")
    buffer = encode.fig_to_img_buffer(fig, fig_type=fig_type, img_format=img_format)
    mime_type = get_mime_type(img_format)
    img_data = encode.buffer_to_base64(buffer, mime_type=mime_type)
    img_tag = f'<img src="{img_data}" download="image.{img_format}"/>'
    download_tags = ""
    for download_format in download_formats:
        buffer = encode.fig_to_img_buffer(fig, fig_type=fig_type, img_format=download_format)
        mime_type = get_mime_type(download_format)
        download_data = encode.buffer_to_base64(buffer, mime_type)
        download_tag = f'<a class="dropdown-item" href="{download_data}" download="image.{download_format}">Download As {download_format.upper()}</a>\n'
        download_tags += download_tag

    # Use grainalyzer-specific endpoints and containers if needed
    if is_grainalyzer:
        delete_endpoint = f"/projects/{project_id}/multivariate/outputs/delete/{output_id}"
        target_container = "#multivariate_outputs"
    else:
        delete_endpoint = f"/projects/{project_id}/outputs/delete/{output_id}"
        target_container = "#outputs_container"

    html = f"""
        <div>
            {img_tag}
            <form method="delete" action="{delete_endpoint}">
                <div class="dropdown show">
                    <a class="btn btn-secondary dropdown-toggle" href="#" role="button" id="{output_id}_dropdown" data-bs-toggle="dropdown" aria-expanded="false">
                        Actions
                    </a>
                    <div class="dropdown-menu" aria-labelledby="{output_id}_dropdown">
                        {download_tags}
                        <button class="dropdown-item" type="submit" data-hx-post="{delete_endpoint}" data-hx-target="{target_container}" data-hx-swap="innerHTML" onclick="show_delete_output_spinner();">Delete This Output</button>
                    </div>
                </div>
            </form>
        </div>
        <hr>"""
    return html


def embed_tabbed_graphs(tabs, output_id, project_id, fig_type="matplotlib", img_format='svg', download_formats=['svg'], is_grainalyzer=False):
    """
    Embed multiple graphs in a Bootstrap tabbed interface.

    Args:
        tabs: List of dicts with 'name' (tab label) and 'fig' (matplotlib figure)
              Example: [{"name": "Graph 1", "fig": fig1}, {"name": "Graph 2", "fig": fig2}]
        output_id: Unique identifier for this output
        project_id: Project ID
        fig_type: Figure type (default: "matplotlib")
        img_format: Display format (default: 'svg')
        download_formats: List of formats for download options
        is_grainalyzer: Whether this is a grainalyzer output (affects delete endpoint)

    Returns:
        HTML string with tabbed interface
    """
    accepted_image_formats = ['jpg', 'jpeg', 'png', 'pdf', 'eps', 'svg']
    if not formats.check(file_format=img_format, accepted_formats=accepted_image_formats):
        raise ValueError("Image format is not supported.")
    for download_format in download_formats:
        if not formats.check(file_format=download_format, accepted_formats=accepted_image_formats):
            raise ValueError("Download format is not supported.")

    if not tabs or len(tabs) == 0:
        raise ValueError("At least one tab is required.")

    # Use grainalyzer-specific endpoints and containers if needed
    if is_grainalyzer:
        delete_endpoint = f"/projects/{project_id}/multivariate/outputs/delete/{output_id}"
        target_container = "#multivariate_outputs"
    else:
        delete_endpoint = f"/projects/{project_id}/outputs/delete/{output_id}"
        target_container = "#outputs_container"

    # Generate unique IDs for tabs (prefix with 'tab_' to ensure valid CSS selectors)
    tab_ids = [f"tab_{output_id}_{i}" for i in range(len(tabs))]

    # Build tab navigation
    tab_nav_items = []
    for i, tab in enumerate(tabs):
        active_class = "active" if i == 0 else ""
        tab_nav_items.append(f'''
            <li class="nav-item" role="presentation">
                <button class="nav-link {active_class}" id="{tab_ids[i]}_nav" data-bs-toggle="tab" data-bs-target="#{tab_ids[i]}" type="button" role="tab" aria-controls="{tab_ids[i]}" aria-selected="{'true' if i == 0 else 'false'}">
                    {tab['name']}
                </button>
            </li>''')

    tab_nav_html = f'''
        <ul class="nav nav-tabs" id="tabs_{output_id}" role="tablist">
            {''.join(tab_nav_items)}
        </ul>'''

    # Build tab content panes
    tab_panes = []
    for i, tab in enumerate(tabs):
        active_class = "show active" if i == 0 else ""

        # Generate image for this tab
        buffer = encode.fig_to_img_buffer(tab['fig'], fig_type=fig_type, img_format=img_format)
        mime_type = get_mime_type(img_format)
        img_data = encode.buffer_to_base64(buffer, mime_type=mime_type)

        tab_panes.append(f'''
            <div class="tab-pane {active_class}" id="{tab_ids[i]}" role="tabpanel" aria-labelledby="{tab_ids[i]}_nav">
                <div style="width: 100%; overflow: auto;">
                    <img src="{img_data}" style="max-width: 100%; height: auto; display: block; margin: 10px 0;"/>
                </div>
            </div>''')

    tab_content_html = f'''
        <div class="tab-content" id="tabContent_{output_id}" style="padding: 15px; min-height: 400px; display: block;">
            {''.join(tab_panes)}
        </div>'''

    # Build download options - create ZIP files bundling all tabs
    download_tags = []
    for download_format in download_formats:
        import zipfile
        import io

        # Create ZIP file in memory
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for i, tab in enumerate(tabs):
                # Generate image for this tab
                img_buffer = encode.fig_to_img_buffer(tab['fig'], fig_type=fig_type, img_format=download_format)
                safe_tab_name = safe_filename(tab['name'])
                filename = f"{safe_tab_name}.{download_format}"

                # Add to ZIP
                zip_file.writestr(filename, img_buffer.getvalue())

        # Encode ZIP as base64
        zip_buffer.seek(0)
        mime_type = 'application/zip'
        download_data = encode.buffer_to_base64(zip_buffer, mime_type)

        # Determine if format is vector or raster
        format_type = "Vectors" if download_format in ['svg', 'pdf', 'eps'] else "Rasters"

        download_tags.append(
            f'<a class="dropdown-item" href="{download_data}" download="charts.zip">Download as ZIP of {format_type} ({download_format.upper()})</a>\n'
        )

    download_tags_html = ''.join(download_tags)

    # Build complete HTML (without inline script - will be initialized by global handler)
    html = f"""
        <div class="tabbed-output-container" data-tabbed-output-id="{output_id}">
            {tab_nav_html}
            {tab_content_html}
            <form method="delete" action="{delete_endpoint}">
                <div class="dropdown show" style="margin-top: 10px;">
                    <a class="btn btn-secondary dropdown-toggle" href="#" role="button" id="dropdown_{output_id}" data-bs-toggle="dropdown" aria-expanded="false">
                        Actions
                    </a>
                    <div class="dropdown-menu" aria-labelledby="dropdown_{output_id}">
                        {download_tags_html}
                        <div class="dropdown-divider"></div>
                        <button class="dropdown-item" type="submit" data-hx-post="{delete_endpoint}" data-hx-target="{target_container}" data-hx-swap="innerHTML" onclick="show_delete_output_spinner();">Delete This Output</button>
                    </div>
                </div>
            </form>
        </div>
        <hr>"""

    return html


def embed_matrix(dataframe: DataFrame, output_id, project_id, title:str = None, download_formats=['xlsx'], is_grainalyzer=False, show_index=True):
    accepted_image_formats = ['xlsx', 'xls', 'csv']
    download_tags = ""
    for download_format in download_formats:
        if download_format not in accepted_image_formats:
            raise ValueError(f"Download format '{download_format}' is not supported.")
    for download_format in download_formats:
        if download_format == 'xlsx':
            buffer = matrices.to_xlsx(dataframe)
        elif download_format == 'xls':
            buffer = matrices.to_xls(dataframe)
        elif download_format == 'csv':
            buffer = matrices.to_csv(dataframe)
        mime_type = get_mime_type(download_format)
        download_data = encode.buffer_to_base64(buffer, mime_type)
        download_tag = f'<a class="dropdown-item" href="{download_data}" download="{safe_filename(title if title else "data")}.{download_format}">Download As {download_format.upper()}</a>\n'
        download_tags += download_tag

    if is_grainalyzer:
        delete_endpoint = f"/projects/{project_id}/multivariate/outputs/delete/{output_id}"
        target_container = "#multivariate_outputs"
    else:
        delete_endpoint = f"/projects/{project_id}/outputs/delete/{output_id}"
        target_container = "#outputs_container"

    if show_index:
        table_html = matrices.dataframe_to_html(dataframe, title=title)
    else:
        table_html = (
            (f"<h4>{title}</h4>" if title else "") +
            dataframe.to_html(
                classes="table table-bordered table-striped",
                justify="center",
                index=False,
            ).replace('<th>', '<th style="background-color: White;">')
             .replace('<td>', '<td style="background-color: White;">')
        )

    html = f"""
        <div>
            {table_html}
            <form method="post" action="{delete_endpoint}">
                <div class="dropdown show">
                    <a class="btn btn-secondary dropdown-toggle" href="#" role="button" id="{output_id}_dropdown" data-bs-toggle="dropdown" aria-expanded="false">
                        Actions
                    </a>
                    <div class="dropdown-menu" aria-labelledby="{output_id}_dropdown">
                        {download_tags}
                        <button class="dropdown-item" type="submit" data-hx-post="{delete_endpoint}" data-hx-target="{target_container}" data-hx-swap="innerHTML" onclick="show_delete_output_spinner();">Delete This Output</button>
                    </div>
                </div>
            </form>
        </div>
        <hr>"""
    return html


def embed_tabbed_matrices(tabs, output_id, project_id, download_formats=['xlsx'], is_grainalyzer=False):
    """
    Embed multiple matrices (DataFrames) in a Bootstrap tabbed interface.

    Args:
        tabs: List of dicts with 'name' (tab label) and 'dataframe' (pandas DataFrame)
              Example: [{"name": "Sample 1", "dataframe": df1}, {"name": "Sample 2", "dataframe": df2}]
        output_id: Unique identifier for this output
        project_id: Project ID
        download_formats: List of formats for download options (default: ['xlsx'])
        is_grainalyzer: Whether this is a grainalyzer output (affects delete endpoint)

    Returns:
        HTML string with tabbed interface
    """
    accepted_formats = ['xlsx', 'xls', 'csv']
    for download_format in download_formats:
        if download_format not in accepted_formats:
            raise ValueError(f"Download format '{download_format}' is not supported.")

    if not tabs or len(tabs) == 0:
        raise ValueError("At least one tab is required.")

    # Use grainalyzer-specific endpoints and containers if needed
    if is_grainalyzer:
        delete_endpoint = f"/projects/{project_id}/multivariate/outputs/delete/{output_id}"
        target_container = "#multivariate_outputs"
    else:
        delete_endpoint = f"/projects/{project_id}/outputs/delete/{output_id}"
        target_container = "#outputs_container"

    # Generate unique IDs for tabs
    tab_ids = [f"tab_{output_id}_{i}" for i in range(len(tabs))]

    # Build tab navigation
    tab_nav_items = []
    for i, tab in enumerate(tabs):
        active_class = "active" if i == 0 else ""
        tab_nav_items.append(f'''
            <li class="nav-item" role="presentation">
                <button class="nav-link {active_class}" id="{tab_ids[i]}_nav" data-bs-toggle="tab" data-bs-target="#{tab_ids[i]}" type="button" role="tab" aria-controls="{tab_ids[i]}" aria-selected="{'true' if i == 0 else 'false'}">
                    {tab['name']}
                </button>
            </li>''')

    tab_nav_html = f'''
        <ul class="nav nav-tabs" id="tabs_{output_id}" role="tablist">
            {''.join(tab_nav_items)}
        </ul>'''

    # Build tab content panes
    tab_panes = []
    for i, tab in enumerate(tabs):
        active_class = "show active" if i == 0 else ""

        # Generate HTML table for this tab
        table_html = matrices.dataframe_to_html(tab['dataframe'], title=None)

        tab_panes.append(f'''
            <div class="tab-pane {active_class}" id="{tab_ids[i]}" role="tabpanel" aria-labelledby="{tab_ids[i]}_nav">
                <div style="width: 100%; overflow: auto; max-height: 600px;">
                    {table_html}
                </div>
            </div>''')

    tab_content_html = f'''
        <div class="tab-content" id="tabContent_{output_id}" style="padding: 15px; min-height: 200px; display: block;">
            {''.join(tab_panes)}
        </div>'''

    # Build download options - bundle tabs intelligently
    download_tags = []
    for download_format in download_formats:
        if download_format in ['xlsx', 'xls']:
            # For Excel formats: Create single file with multiple sheets
            import openpyxl
            import io

            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            for tab in tabs:
                safe_sheet_name = safe_filename(tab['name'])[:31]  # Excel sheet name limit
                ws = wb.create_sheet(title=safe_sheet_name)

                # Write dataframe to sheet
                df = tab['dataframe']
                # Write headers
                for col_idx, col_name in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                # Write data
                for row_idx, row_data in enumerate(df.values, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            mime_type = get_mime_type(download_format)
            download_data = encode.buffer_to_base64(buffer, mime_type)
            download_tags.append(
                f'<a class="dropdown-item" href="{download_data}" download="data.{download_format}">Download as Spreadsheet ({download_format.upper()})</a>\n'
            )

        elif download_format == 'csv':
            # For CSV: Create ZIP file with individual CSV files
            import zipfile
            import io

            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for tab in tabs:
                    csv_buffer = matrices.to_csv(tab['dataframe'])
                    safe_tab_name = safe_filename(tab['name'])
                    filename = f"{safe_tab_name}.csv"
                    zip_file.writestr(filename, csv_buffer.getvalue())

            zip_buffer.seek(0)
            mime_type = 'application/zip'
            download_data = encode.buffer_to_base64(zip_buffer, mime_type)
            download_tags.append(
                f'<a class="dropdown-item" href="{download_data}" download="data.zip">Download as ZIP of Spreadsheets (CSV)</a>\n'
            )

    download_tags_html = ''.join(download_tags)

    # Build complete HTML
    html = f"""
        <div class="tabbed-output-container" data-tabbed-output-id="{output_id}">
            {tab_nav_html}
            {tab_content_html}
            <form method="delete" action="{delete_endpoint}">
                <div class="dropdown show" style="margin-top: 10px;">
                    <a class="btn btn-secondary dropdown-toggle" href="#" role="button" id="dropdown_{output_id}" data-bs-toggle="dropdown" aria-expanded="false">
                        Actions
                    </a>
                    <div class="dropdown-menu" aria-labelledby="dropdown_{output_id}">
                        {download_tags_html}
                        <div class="dropdown-divider"></div>
                        <button class="dropdown-item" type="submit" data-hx-post="{delete_endpoint}" data-hx-target="{target_container}" data-hx-swap="innerHTML" onclick="show_delete_output_spinner();">Delete This Output</button>
                    </div>
                </div>
            </form>
        </div>
        <hr>"""

    return html
